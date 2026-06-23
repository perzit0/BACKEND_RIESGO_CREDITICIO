import io
from collections import Counter
from flask import Blueprint, jsonify, Response
from extensions import db
from models.usuario import Usuario
from models.evaluacion import Evaluacion
from services.middleware import requiere_rol

admin_bp = Blueprint("admin", __name__)


def _analizar_dni(dni: str) -> list:
    """Detecta patrones sospechosos en un DNI. Devuelve lista de razones."""
    razones = []

    if not (len(dni) == 8 and dni.isdigit()):
        razones.append("DNI con formato inválido (debe tener 8 dígitos numéricos)")
        return razones

    digitos = [int(d) for d in dni]

    # 1. Todos los dígitos iguales
    if len(set(dni)) == 1:
        razones.append(f"DNI con todos los dígitos iguales ({dni})")

    # 2. Secuencia ascendente o descendente
    diffs = [digitos[i+1] - digitos[i] for i in range(len(digitos)-1)]
    if all(d == 1 for d in diffs):
        razones.append(f"DNI con secuencia ascendente ({dni})")
    elif all(d == -1 for d in diffs):
        razones.append(f"DNI con secuencia descendente ({dni})")

    # 3. Pares de dígitos repetidos (77, 88, 55...)
    pares = [dni[i:i+2] for i in range(0, 8, 2)]
    pares_dobles = [p for p in pares if p[0] == p[1]]
    if len(pares_dobles) >= 2:
        razones.append(f"DNI con múltiples pares repetidos: {', '.join(pares_dobles)} ({dni})")

    # 4. Patrón alternado (79797979, 12121212)
    if dni[::2] == dni[0] * 4 and dni[1::2] == dni[1] * 4:
        razones.append(f"DNI con patrón alternado ({dni[0]}{dni[1]} repetido)")

    # 5. Un dígito aparece 6 o más veces
    conteo = Counter(dni)
    digito_comun, frecuencia = conteo.most_common(1)[0]
    if frecuencia >= 6:
        razones.append(f"DNI con dígito '{digito_comun}' repetido {frecuencia} veces ({dni})")

    # 6. Primera mitad igual a segunda mitad
    if dni[:4] == dni[4:]:
        razones.append(f"DNI con primera mitad igual a segunda mitad ({dni[:4]}-{dni[4:]})")

    return razones


@admin_bp.route("/dashboard", methods=["GET"])
@requiere_rol("admin")
def dashboard():
    total_usuarios = Usuario.query.filter_by(rol="user").count()
    total_evaluaciones = Evaluacion.query.count()

    distribucion = {"bajo": 0, "medio": 0, "alto": 0}
    for categoria, _ in (
        Evaluacion.query
        .with_entities(Evaluacion.categoria_riesgo, Evaluacion.id)
        .all()
    ):
        if categoria in distribucion:
            distribucion[categoria] += 1

    return jsonify({
        "total_usuarios": total_usuarios,
        "total_evaluaciones": total_evaluaciones,
        "distribucion_riesgo": distribucion,
    }), 200


@admin_bp.route("/usuarios", methods=["GET"])
@requiere_rol("admin")
def listar_usuarios():
    usuarios = Usuario.query.filter_by(rol="user").all()
    resultado = [
        {
            "id": u.id,
            "nombre": u.nombre,
            "email": u.email,
            "correo_verificado": u.correo_verificado,
            "telefono_verificado": u.telefono_verificado,
            "marcado_fraude": u.marcado_fraude,
            "bloqueado": u.bloqueado,
            "fecha_registro": u.fecha_registro.isoformat(),
        }
        for u in usuarios
    ]
    return jsonify(resultado), 200


@admin_bp.route("/casos-fraude", methods=["GET"])
@requiere_rol("admin")
def casos_fraude():
    sospechosos = Usuario.query.filter_by(
        marcado_fraude=True, revisado_por_admin=False, bloqueado=False
    ).all()

    resultado = []
    for u in sospechosos:
        razones = []

        # 1. Analizar DNI
        razones_dni = _analizar_dni(u.dni or "")
        razones.extend(razones_dni)

        # 2. IP con múltiples registros recientes
        from datetime import datetime, timedelta
        hace_10_min = datetime.utcnow() - timedelta(minutes=10)
        registros_ip = Usuario.query.filter(
            Usuario.ip_registro == u.ip_registro,
            Usuario.fecha_registro >= hace_10_min,
            Usuario.id != u.id
        ).count()
        if registros_ip > 0:
            razones.append(f"IP {u.ip_registro} usada en {registros_ip} registro(s) reciente(s)")

        # 3. Hora de registro sospechosa (1am - 5am)
        hora = u.fecha_registro.hour if u.fecha_registro else 12
        if 1 <= hora <= 5:
            razones.append(f"Registro a las {hora:02d}:00 (horario inusual)")

        # 4. Teléfono repetido en más de 2 cuentas
        otros_con_mismo_tel = Usuario.query.filter(
            Usuario.telefono == u.telefono,
            Usuario.id != u.id
        ).count()
        if otros_con_mismo_tel > 2:
            razones.append(f"Teléfono registrado en más de 2 cuentas distintas")

        if not razones:
            razones.append("Patrón de comportamiento detectado por el modelo de IA")

        resultado.append({
            "id": u.id,
            "nombre": u.nombre,
            "dni": u.dni,
            "email": u.email,
            "telefono": u.telefono,
            "ip_registro": u.ip_registro,
            "fecha_registro": u.fecha_registro.isoformat(),
            "razones_fraude": razones,
        })

    return jsonify(resultado), 200


@admin_bp.route("/casos-fraude/<int:usuario_id>/resolver", methods=["POST"])
@requiere_rol("admin")
def resolver_caso_fraude(usuario_id):
    usuario = db.session.get(Usuario, usuario_id)
    if usuario is None:
        return jsonify({"error": "Usuario no encontrado"}), 404
    usuario.revisado_por_admin = True
    usuario.bloqueado = False
    db.session.commit()
    return jsonify({"mensaje": f"Cuenta de {usuario.email} aprobada correctamente"}), 200


@admin_bp.route("/casos-fraude/<int:usuario_id>/bloquear", methods=["POST"])
@requiere_rol("admin")
def bloquear_caso_fraude(usuario_id):
    usuario = db.session.get(Usuario, usuario_id)
    if usuario is None:
        return jsonify({"error": "Usuario no encontrado"}), 404
    usuario.bloqueado = True
    usuario.revisado_por_admin = True
    db.session.commit()
    return jsonify({"mensaje": f"Cuenta de {usuario.email} bloqueada correctamente"}), 200


@admin_bp.route("/metricas-modelos", methods=["GET"])
@requiere_rol("admin")
def metricas_modelos():
    return jsonify({
        "comparativa_riesgo": {
            "xgboost": {
                "accuracy": 0.8731, "f1_score": 0.6435, "roc_auc": 0.8716,
                "tiempo_entrenamiento_seg": 4.1195, "peso_kb": 321.61
            },
            "random_forest": {
                "accuracy": 0.8638, "f1_score": 0.6065, "roc_auc": 0.8448,
                "tiempo_entrenamiento_seg": 6.9632, "peso_kb": 54192.39
            },
            "logistic_regression": {
                "accuracy": 0.8390, "f1_score": 0.5132, "roc_auc": 0.7928,
                "tiempo_entrenamiento_seg": 0.3372, "peso_kb": 3.95
            },
            "gradient_boosting": {
                "accuracy": 0.8685, "f1_score": 0.5993, "roc_auc": 0.8595,
                "tiempo_entrenamiento_seg": 7.5636, "peso_kb": 141.61
            },
            "ganador": "xgboost",
        },
        "comparativa_fraude": {
            "isolation_forest": {
                "tasa_deteccion": 0.315, "falsos_positivos": 0.0412,
                "tiempo_entrenamiento_seg": 0.1902, "peso_kb": 1533.07
            },
            "random_forest_supervisado": {
                "tasa_deteccion": 0.99, "falsos_positivos": 0.0,
                "tiempo_entrenamiento_seg": 0.2027, "peso_kb": 810.21
            },
            "ganador": "random_forest_supervisado",
        },
        "modelo_riesgo_produccion": {
            "algoritmo": "XGBoost",
            "accuracy": 0.8731, "f1_score": 0.6435,
            "roc_auc": 0.8716, "tiempo_entrenamiento_seg": 4.1195, "peso_kb": 321.61,
        },
        "modelo_fraude_produccion": {
            "algoritmo": "Random Forest Supervisado",
            "tasa_deteccion": 0.99, "falsos_positivos": 0.0,
            "tiempo_entrenamiento_seg": 0.2027, "peso_kb": 810.21,
        },
    }), 200


@admin_bp.route("/exportar-excel", methods=["GET"])
@requiere_rol("admin")
def exportar_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl no instalado. Ejecuta: pip install openpyxl"}), 500

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios y Evaluaciones"

    color_header = "6B4EFF"
    color_bajo = "D1FAE5"
    color_medio = "FEF3C7"
    color_alto = "FEE2E2"
    color_fila_par = "F5F3FF"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=color_header)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_side = Side(style="thin", color="E2E8F0")
    cell_border = Border(
        left=border_side, right=border_side,
        top=border_side, bottom=border_side
    )

    columnas = [
        "ID", "Nombre", "DNI", "Correo", "Telefono",
        "Correo verificado", "Telefono verificado",
        "Marcado fraude", "Bloqueado", "Fecha registro",
        "Ultima evaluacion", "Nivel de riesgo",
        "Ingreso mensual (S/.)", "Monto en bancos (S/.)",
        "Num. cuentas", "Creditos previos", "Dias de mora",
        "Edad", "Antiguedad laboral (meses)",
        "Tipo empleo", "Nivel educativo", "Estado civil", "Tipo vivienda",
    ]

    ws.row_dimensions[1].height = 35
    for col_idx, col_name in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = cell_border

    from sqlalchemy import func

    subq = (
        db.session.query(
            Evaluacion.usuario_id,
            func.max(Evaluacion.fecha_evaluacion).label("ultima_fecha")
        )
        .group_by(Evaluacion.usuario_id)
        .subquery()
    )

    ultimas_evaluaciones = (
        db.session.query(Evaluacion)
        .join(subq, (Evaluacion.usuario_id == subq.c.usuario_id) &
              (Evaluacion.fecha_evaluacion == subq.c.ultima_fecha))
        .all()
    )

    ultima_por_usuario = {e.usuario_id: e for e in ultimas_evaluaciones}
    usuarios = Usuario.query.filter_by(rol="user").order_by(Usuario.fecha_registro.desc()).all()

    def _cat_fill(cat):
        if cat == "bajo":   return PatternFill("solid", fgColor=color_bajo)
        if cat == "medio":  return PatternFill("solid", fgColor=color_medio)
        if cat == "alto":   return PatternFill("solid", fgColor=color_alto)
        return None

    SIN_DATO = "S/D"

    for row_idx, u in enumerate(usuarios, 2):
        ultima = ultima_por_usuario.get(u.id)
        es_par = row_idx % 2 == 0
        fila_fill = PatternFill("solid", fgColor=color_fila_par) if es_par else None

        valores = [
            u.id, u.nombre, u.dni, u.email, u.telefono,
            "Si" if u.correo_verificado else "No",
            "Si" if u.telefono_verificado else "No",
            "Si" if u.marcado_fraude else "No",
            "Si" if u.bloqueado else "No",
            u.fecha_registro.strftime("%d/%m/%Y %H:%M") if u.fecha_registro else "",
            ultima.fecha_evaluacion.strftime("%d/%m/%Y %H:%M") if ultima else "Sin evaluacion",
            ultima.categoria_riesgo.upper() if ultima else SIN_DATO,
            ultima.ingreso_mensual if ultima else SIN_DATO,
            ultima.monto_en_bancos if ultima else SIN_DATO,
            ultima.num_cuentas_bancarias if ultima else SIN_DATO,
            ultima.num_creditos_previos if ultima else SIN_DATO,
            ultima.dias_mora_historico if ultima else SIN_DATO,
            ultima.edad if ultima else SIN_DATO,
            ultima.antiguedad_laboral_meses if ultima else SIN_DATO,
            ultima.tipo_empleo or SIN_DATO if ultima else SIN_DATO,
            ultima.nivel_educativo or SIN_DATO if ultima else SIN_DATO,
            ultima.estado_civil or SIN_DATO if ultima else SIN_DATO,
            ultima.tipo_vivienda or SIN_DATO if ultima else SIN_DATO,
        ]

        for col_idx, valor in enumerate(valores, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center", horizontal="center")
            if col_idx == 12 and ultima:
                fill = _cat_fill(ultima.categoria_riesgo)
                if fill:
                    cell.fill = fill
                    cell.font = Font(bold=True)
            elif fila_fill:
                cell.fill = fila_fill

        ws.row_dimensions[row_idx].height = 22

    anchos = [
        6, 22, 12, 30, 16, 14, 16,
        14, 12, 18, 18, 14,
        16, 18, 12, 14, 12,
        8, 20, 14, 16, 14, 14,
    ]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=usuarios_riesgo_crediticio.xlsx"
        }
    )